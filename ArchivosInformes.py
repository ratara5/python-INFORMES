#Este programa muestra la tarjeta de año teocrático de publicador
#a partir del libro con hojas de informes por mes
#se pueden hacer ajustes (con el código existente) para mostrar 
#la tarjeta por periodos menores al año

import openpyxl as xl
import pandas as pd
import numpy as np
from pandas.core.indexes.base import Index

############################
# DEFINIR DATOS DE ENTRADA #
############################

# Borrar hojas no usadas en libros 
def borrar_hojas(year,group):
    fichero='E:/TEOCRÁTICO/MACHADO 14944/Informes/'+str(year)+'/Informes/Grupo #'+str(group)+'.xlsx'
    lb = xl.load_workbook(filename=fichero)
    for j in range(8): 
        lb.remove(lb[lb.sheetnames[1]])
    lb.save(fichero)
    print('quedaron estas hojas del grupo '+str(group)+':')
    print (lb.sheetnames)

# Crear lista de meses por semestre de año
def list_mes_semes(semester,year):
    if semester=="1":
        list=["Septiembre ", "Octubre ","Noviembre ","Diciembre ","Enero","Febrero"]
    else:
        list=["Marzo ", "Abril ","Mayo ","Junio ","Julio ","Agosto "]
    for i in range(0,len(list)):
        list[i]+=year
    return list

# Crear lista de meses por año
def list_mes_anio(year):
    list=["Septiembre ", "Octubre ","Noviembre ","Diciembre ","Enero ","Febrero ","Marzo ", "Abril ","Mayo ","Junio ","Julio ","Agosto "]
    for i in range(0,4):
        list[i]+=str(int(year)-1)
    for i in range(4,12):
        list[i]+=year
    return list

# Agregar encabezado superior a cada dataframe
def enc_sup_mes (df,mes): 
    df.columns= pd.MultiIndex.from_tuples(zip(['Nombre', mes, mes, mes, mes, mes, mes], df.columns))
    return df

# Combinar todos dataframes por la izquierda para nombres coincidentes (llave 'Nombre')
def mergear(marco):
    for i in range(1,6):
        marco=pd.merge(left=marco, right=globals()['archivoMes'+str(i)], how='left',on=('Nombre','Nombre'))
    return marco


anio=input(str('Escriba año teocrático: '))
listaMeses=list_mes_anio(anio)
print("Estos son los meses del año teocrático "+str(anio)+" :   "+str(listaMeses))

###############################
# CREAR DATAFRAME DE CADA MES #
###############################

lis_fil_era=[0,1,32,33,34,45] 
nom_col=['Nombre','Publicaciones','Videos','Horas','Revisitas','Cursos bíblicos','Observaciones']

grupo=input(str('Elija el grupo. Del 1 al 9: '))

# Formar dataframes por mes del grupo
for i in range (0,12):

    globals()["archivoMes"+str(i)]=pd.read_excel('E:/TEOCRÁTICO/MACHADO 14944/Informes/2021/Informes/Grupo #'+grupo+'.xlsx',sheet_name=listaMeses[i],header=None,names=nom_col, skiprows=lis_fil_era,usecols = "B,D:H,K").dropna(how='all')

    globals()["archivoMes"+str(i)].index=globals()["archivoMes"+str(i)]['Nombre'] 
    globals()["archivoMes"+str(i)]=globals()["archivoMes"+str(i)].drop(['Nombre'],axis=1) 
   
    # print("Miembros Grupo # "+grupo+" Mes "+listaMeses[i])
    # print()
    # #print(globals()["archivoMes"+str(i)]) # Mostrar dataframes creados
    # print("------------------------------")

######################################
# CREAR DATAFRAME DE CADA PUBLICADOR #
######################################

def marco_publicador(nom_pub):
    marcoPublicador=pd.DataFrame([])
    for i in range(0,12):
        if nombre in globals()["archivoMes"+str(i)].index:
            marcoPublicador=marcoPublicador.append(globals()["archivoMes"+str(i)].loc[nom_pub])
            marcoPublicador=marcoPublicador.rename(index={nombre:listaMeses[i]})
            #marcoPublicador['Horas']=(marcoPublicador['Horas']-marcoPublicador['Horas'].astype(int))*100/60+marcoPublicador['Horas'].astype(int) # para hh:mm
        else:
            marcoPublicador=marcoPublicador.append([])
            marcoPublicador=marcoPublicador.rename(index={nombre:listaMeses[i]})
    return marcoPublicador
    #ERROR SI: Publicador no conserva nombre pa' cada mes // Publicador no existe desde mes inicial de listaMeses OPTIMIZAR ESTA PARTE CON UN CICLO


def fila_totales(marco):
    marco.loc['Totales']=np.around(marco.iloc[:,0:5].sum(),2)
    return marco

def fila_promedios(marco):
    filas=marco.shape[0]-1
    marco.loc['Promedios']=np.around(marco.iloc[0:filas,0:5].mean(),2)
    return marco

lis_nom=[]
for nombre in archivoMes11.index:
    #if nombre in archivoMes0.index and nombre in archivoMes5.index: 
    globals()["marcoPublicador"+nombre]=marco_publicador(nombre)
    globals()["marcoPublicador"+nombre]=fila_totales(globals()["marcoPublicador"+nombre])
    globals()["marcoPublicador"+nombre]=fila_promedios(globals()["marcoPublicador"+nombre])
    
    print('')
    print('')
    print('TARJETA '+nombre)
    print('Información del Grupo #'+grupo)
    print('')
    
    print(globals()["marcoPublicador"+nombre])
    lis_nom.append(nombre)
        
###############################################
# TOTALES Y PROMEDIOS INDIVIDUALES CADA GRUPO #
###############################################

# Formar dataframes totales del grupo
marcoTotales=pd.DataFrame([])
for nombre in lis_nom:
    marcoTotales=marcoTotales.append(globals()["marcoPublicador"+nombre].loc['Totales'])
    marcoTotales = marcoTotales.rename(index={'Totales':nombre})

print('')
print('-------------------------')
print('Esta es la tabla de TOTALES del Grupo #'+grupo+' para el año de servicio '+anio)
print('')
print(marcoTotales)

# Formar dataframes promedios del grupo
marcoPromedios=pd.DataFrame([])
for nombre in lis_nom:
    marcoPromedios=marcoPromedios.append(globals()["marcoPublicador"+nombre].loc['Promedios'])
    marcoPromedios = marcoPromedios.rename(index={'Promedios':nombre})

print('')
print('-------------------------')
print('Esta es la tabla de PROMEDIOS del Grupo #'+grupo+' para el año de servicio '+anio)
print('')
print(marcoPromedios)
print('')
print(marcoPromedios.count(axis=0))

# Código averiguar en qué grupos ha estado publicador en un semestre

marcoPromedios.to_excel('E:/TEOCRÁTICO/MACHADO 14944/Informes/2021/Informes/Grupo # Promedios.xlsx',sheet_name='Promedios')
marcoTotales.to_excel('E:/TEOCRÁTICO/MACHADO 14944/Informes/2021/Informes/Grupo # Totales.xlsx',sheet_name='Totales')
